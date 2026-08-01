"""F7-T3 — tests for the batch result writer.

Everything runs against tmp_path files built in-test: no network, no fixture
data on disk, no geocoder. The tests that matter most here are the two that
protect people rather than data shape — the formula-injection test (plan D21)
and the partial-write test (plan R15).
"""
import csv

import pytest

from app.batch import (
    STATUS_ERROR,
    STATUS_MATCHED,
    STATUS_NO_GEOCODE,
    STATUS_OUTSIDE,
    BatchError,
    RowResult,
    result_columns,
)
from app.batch.writer import neutralize_cell, resolve_format, write_results

# An operator's file, in an order nothing may disturb: the id column is last on
# purpose, and "score" collides in spirit with pip_score.
HEADERS = ("address", "score", "case_id")
ATTRIBUTES = ("dist_num", "dist_label")

# The classic DDE payload. If this survives to a cell unescaped, opening the
# output in Excel prompts to launch a program.
INJECTION = "=cmd|'/c calc'!A1"


def _matched(address="121 N La Salle St", **overrides) -> RowResult:
    fields = dict(
        row={"address": address, "score": "manual-7", "case_id": "C-001"},
        status=STATUS_MATCHED,
        matched_address="121 N LA SALLE ST, CHICAGO 60602",
        score=100.0,
        provider="local_offline",
        lon=-87.63192,
        lat=41.88354,
        feature={"dist_num": "17", "dist_label": "Albany Park"},
    )
    fields.update(overrides)
    return RowResult(**fields)


def _read_csv(path) -> tuple[list[str], list[list[str]]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))
    return rows[0], rows[1:]


def _read_xlsx(path) -> list[list]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True)
    try:
        return [list(row) for row in workbook.active.iter_rows(values_only=True)]
    finally:
        workbook.close()


# ---- format resolution ----

def test_format_inferred_from_suffix(tmp_path):
    assert resolve_format(tmp_path / "out.csv") == "csv"
    assert resolve_format(tmp_path / "OUT.CSV") == "csv"
    assert resolve_format(tmp_path / "out.xlsx") == "xlsx"


def test_explicit_format_overrides_suffix(tmp_path):
    assert resolve_format(tmp_path / "out.xlsx", "csv") == "csv"
    assert resolve_format(tmp_path / "out.dat", ".CSV") == "csv"


def test_unknown_suffix_raises_batch_error(tmp_path):
    with pytest.raises(BatchError, match="cannot infer output format"):
        resolve_format(tmp_path / "results.xls")


def test_unsupported_explicit_format_raises_batch_error(tmp_path):
    with pytest.raises(BatchError, match="unsupported output format"):
        resolve_format(tmp_path / "out.csv", "parquet")


def test_write_results_rejects_unknown_suffix(tmp_path):
    target = tmp_path / "results.txt"
    with pytest.raises(BatchError):
        write_results([_matched()], path=target, headers=HEADERS)
    # Nothing is created for a run that can never succeed.
    assert not target.exists()


# ---- CSV round trip ----

def test_csv_round_trip_preserves_original_columns_then_results(tmp_path):
    target = tmp_path / "out.csv"
    count = write_results(
        [_matched()], path=target, headers=HEADERS, layer_attributes=ATTRIBUTES
    )
    assert count == 1

    header, rows = _read_csv(target)
    assert tuple(header) == HEADERS + result_columns(ATTRIBUTES)
    # The operator's own columns come first, in their original order, unaltered.
    assert header[: len(HEADERS)] == list(HEADERS)
    assert rows[0][: len(HEADERS)] == ["121 N La Salle St", "manual-7", "C-001"]
    # Their "score" column is untouched; ours is the prefixed one.
    assert rows[0][header.index("score")] == "manual-7"
    assert rows[0][header.index("pip_score")] == "100.0"
    assert rows[0][header.index("pip_status")] == STATUS_MATCHED
    assert rows[0][header.index("pip_matched_address")] == (
        "121 N LA SALLE ST, CHICAGO 60602"
    )
    assert rows[0][header.index("pip_provider")] == "local_offline"
    assert rows[0][header.index("pip_lon")] == "-87.63192"
    assert rows[0][header.index("pip_lat")] == "41.88354"
    assert rows[0][header.index("pip_dist_num")] == "17"
    assert rows[0][header.index("pip_dist_label")] == "Albany Park"


def test_csv_returns_row_count_and_writes_every_row(tmp_path):
    target = tmp_path / "out.csv"
    results = [_matched(address=f"{n} Main St") for n in range(5)]
    count = write_results(
        results, path=target, headers=HEADERS, layer_attributes=ATTRIBUTES
    )
    assert count == 5
    _, rows = _read_csv(target)
    assert len(rows) == 5
    assert [row[0] for row in rows] == [f"{n} Main St" for n in range(5)]


def test_empty_run_writes_header_only_and_returns_zero(tmp_path):
    target = tmp_path / "out.csv"
    assert write_results(iter([]), path=target, headers=HEADERS) == 0
    header, rows = _read_csv(target)
    assert tuple(header) == HEADERS + result_columns(())
    assert rows == []


def test_none_result_fields_become_empty_cells_not_the_string_none(tmp_path):
    target = tmp_path / "out.csv"
    write_results(
        [
            RowResult(
                row={"address": "nowhere", "score": "", "case_id": "C-002"},
                status=STATUS_NO_GEOCODE,
                reason="no_candidates",
            )
        ],
        path=target,
        headers=HEADERS,
        layer_attributes=ATTRIBUTES,
    )
    header, rows = _read_csv(target)
    assert rows[0][header.index("pip_status")] == STATUS_NO_GEOCODE
    assert rows[0][header.index("pip_reason")] == "no_candidates"
    for column in ("pip_matched_address", "pip_score", "pip_lon", "pip_dist_num"):
        assert rows[0][header.index(column)] == ""


def test_missing_and_extra_row_keys_follow_the_header_contract(tmp_path):
    # A ragged row must not abort the run or shift columns; a stray key the
    # header does not declare must not sneak an extra column into the output.
    target = tmp_path / "out.csv"
    write_results(
        [
            RowResult(
                row={"address": "1 Main St", "surprise": "should not appear"},
                status=STATUS_OUTSIDE,
                reason="point_outside_all_polygons",
                lon=-87.6,
                lat=41.9,
            )
        ],
        path=target,
        headers=HEADERS,
        layer_attributes=ATTRIBUTES,
    )
    header, rows = _read_csv(target)
    assert len(rows[0]) == len(header)
    assert "should not appear" not in rows[0]
    assert rows[0][: len(HEADERS)] == ["1 Main St", "", ""]
    assert rows[0][header.index("pip_status")] == STATUS_OUTSIDE


def test_missing_layer_attribute_yields_empty_cell(tmp_path):
    target = tmp_path / "out.csv"
    write_results(
        [_matched(feature={"dist_num": "17"})],
        path=target,
        headers=HEADERS,
        layer_attributes=ATTRIBUTES,
    )
    header, rows = _read_csv(target)
    assert rows[0][header.index("pip_dist_num")] == "17"
    assert rows[0][header.index("pip_dist_label")] == ""


# ---- formula injection (plan D21) ----

def test_injection_neutralized_in_matched_address_and_operator_column(tmp_path):
    target = tmp_path / "out.csv"
    write_results(
        [
            RowResult(
                # The operator's own column can carry a payload too — their file
                # may have come from an untrusted upload.
                row={"address": INJECTION, "score": "1", "case_id": "C-003"},
                status=STATUS_MATCHED,
                # ...and this one comes straight from a third-party geocoder.
                matched_address=INJECTION,
                score=99.0,
                provider="arcgis_rest",
                lon=-87.6,
                lat=41.9,
                feature={"dist_num": "17", "dist_label": "Albany Park"},
            )
        ],
        path=target,
        headers=HEADERS,
        layer_attributes=ATTRIBUTES,
    )
    header, rows = _read_csv(target)

    matched_cell = rows[0][header.index("pip_matched_address")]
    operator_cell = rows[0][header.index("address")]
    # BOTH cells are neutralized, and the payload itself is preserved verbatim
    # after the escape so the operator can still see what came back.
    assert matched_cell == "'" + INJECTION
    assert operator_cell == "'" + INJECTION
    assert not matched_cell.startswith("=")
    assert not operator_cell.startswith("=")


@pytest.mark.parametrize(
    "payload", ["=1+1", "+SUM(A1)", "-1+1", "@SUM(A1)", "\tcalc", "\rcalc"]
)
def test_every_trigger_character_is_escaped(tmp_path, payload):
    target = tmp_path / "out.csv"
    write_results(
        [
            RowResult(
                row={"address": "x", "score": "", "case_id": ""},
                status=STATUS_ERROR,
                reason=payload,
            )
        ],
        path=target,
        headers=HEADERS,
    )
    header, rows = _read_csv(target)
    assert rows[0][header.index("pip_reason")] == "'" + payload


@pytest.mark.parametrize(
    "benign",
    ["121 N La Salle St", "", "17", "Albany Park", "O'Hare", " =not-leading", "0.5"],
)
def test_benign_text_is_never_mangled(benign):
    # The escape must be surgical: a general-purpose "sanitizer" that rewrites
    # ordinary addresses would corrupt every row of a real caseload.
    assert neutralize_cell(benign) == benign


def test_negative_coordinate_is_left_as_a_usable_number(tmp_path):
    # pip_lon is negative for every address in Chicago, so it trips the "-"
    # trigger on every row. A cell that is wholly a number cannot be a formula,
    # so it is exempt — otherwise the coordinate column would arrive as text in
    # every output file the service has ever produced.
    target = tmp_path / "out.csv"
    write_results([_matched()], path=target, headers=HEADERS)
    header, rows = _read_csv(target)
    assert rows[0][header.index("pip_lon")] == "-87.63192"
    assert float(rows[0][header.index("pip_lon")]) == pytest.approx(-87.63192)


@pytest.mark.parametrize("number", ["-87.63192", "+41.9", "-1", "-1.0e-5", "-.5"])
def test_plain_numbers_are_exempt_from_escaping(number):
    assert neutralize_cell(number) == number


@pytest.mark.parametrize(
    "not_a_number", ["-1+1", "+1-cmd", "-1 ", "-", "-1a", "=1", "@1", "-1,000"]
)
def test_number_carve_out_does_not_leak_formulas(not_a_number):
    # The exemption must be exact-match only: anything a spreadsheet could still
    # evaluate has to be escaped.
    assert neutralize_cell(not_a_number) == "'" + not_a_number


def test_csv_escapes_a_malicious_header_name(tmp_path):
    # The header row is the operator's input column names, verbatim — as
    # attacker-controlled as any data cell. It used to be written raw, straight
    # past the D21 escape, so an input file whose first column is named
    # "=cmd|'/c calc'!A0" produced an output whose A1 fires on open.
    target = tmp_path / "evil-header.csv"
    write_results(
        [_matched(address="1 Main St")],
        path=target,
        headers=(INJECTION, "score", "case_id"),
    )
    header, rows = _read_csv(target)
    assert header[0] == "'" + INJECTION
    assert not header[0].startswith("=")
    # The data row still lines up with the header it was written under.
    assert len(rows[0]) == len(header)


def test_xlsx_malicious_header_is_not_a_live_formula_cell(tmp_path):
    from openpyxl import load_workbook

    target = tmp_path / "evil-header.xlsx"
    write_results(
        [_matched(address="1 Main St")],
        path=target,
        headers=(INJECTION, "score", "case_id"),
    )

    workbook = load_workbook(target)
    try:
        cell = workbook.active["A1"]
        # 'f' means openpyxl committed a <f> element to sheet1.xml: Excel will
        # evaluate it on open, with no CSV-import warning to stop it.
        assert cell.data_type != "f"
        assert cell.value == "'" + INJECTION
    finally:
        workbook.close()


# ---- re-running over prior output (pip_ column collision) ----

def test_source_that_already_has_pip_columns_is_refused(tmp_path):
    # The obvious second run: take this feature's own output and locate it
    # against a second layer to get wards AND districts on one sheet. Blind
    # concatenation emitted pip_status twice, producing a file our own reader
    # (sources._headers_from) refuses — the pipeline could not round-trip
    # itself. Fail fast instead, and never silently rename.
    target = tmp_path / "second-pass.csv"
    already_located = ("address", "case_id", "pip_status", "pip_lat")

    with pytest.raises(BatchError) as raised:
        write_results([_matched()], path=target, headers=already_located)

    message = str(raised.value)
    assert "duplicate column name(s)" in message
    assert "'pip_status'" in message and "'pip_lat'" in message
    # The operator is told what to do about it.
    assert "new output file" in message
    # And nothing was written for a run that can never succeed.
    assert not target.exists()


def test_xlsx_source_that_already_has_pip_columns_is_refused(tmp_path):
    target = tmp_path / "second-pass.xlsx"
    with pytest.raises(BatchError, match="duplicate column name"):
        write_results(
            [_matched()], path=target, headers=("address", "pip_status")
        )
    assert not target.exists()


def test_layer_attribute_collision_is_refused(tmp_path):
    # The collision is not limited to the fixed columns: a source carrying
    # pip_dist_num from a prior district run collides with this layer too.
    target = tmp_path / "second-pass.csv"
    with pytest.raises(BatchError, match="'pip_dist_num'"):
        write_results(
            [_matched()],
            path=target,
            headers=("address", "pip_dist_num"),
            layer_attributes=ATTRIBUTES,
        )


def test_unprefixed_near_miss_column_is_still_allowed_through(tmp_path):
    # "status" is not "pip_status". The whole point of the prefix is that an
    # operator's own column names stay usable, so the collision check must not
    # over-reach and start rejecting ordinary files.
    target = tmp_path / "near-miss.csv"
    count = write_results(
        [
            RowResult(
                row={"address": "1 Main St", "status": "open", "lat": "n/a"},
                status=STATUS_MATCHED,
                lat=41.9,
            )
        ],
        path=target,
        headers=("address", "status", "lat"),
    )
    assert count == 1
    header, rows = _read_csv(target)
    assert header[:3] == ["address", "status", "lat"]
    # Untouched: their values, not ours.
    assert rows[0][header.index("status")] == "open"
    assert rows[0][header.index("lat")] == "n/a"
    assert rows[0][header.index("pip_status")] == STATUS_MATCHED
    assert rows[0][header.index("pip_lat")] == "41.9"


# ---- partial writes (plan R15) ----

def test_iterator_failure_leaves_a_valid_partial_csv(tmp_path):
    target = tmp_path / "partial.csv"

    def failing_results():
        for n in range(3):
            yield _matched(address=f"{n} Main St")
        raise RuntimeError("provider died mid-run")

    with pytest.raises(RuntimeError, match="provider died mid-run"):
        write_results(
            failing_results(),
            path=target,
            headers=HEADERS,
            layer_attributes=ATTRIBUTES,
        )

    # The file is closed, complete through row 3, and readable by a plain reader.
    header, rows = _read_csv(target)
    assert tuple(header) == HEADERS + result_columns(ATTRIBUTES)
    assert len(rows) == 3
    assert [row[0] for row in rows] == ["0 Main St", "1 Main St", "2 Main St"]
    assert all(len(row) == len(header) for row in rows)


def test_rows_are_on_disk_before_the_iterator_finishes(tmp_path):
    # Proves the write is genuinely incremental rather than buffered to the end:
    # the file already holds row 1 while row 2 is still being produced.
    target = tmp_path / "incremental.csv"
    seen_midway = []

    def probing_results():
        yield _matched(address="first")
        seen_midway.append(target.read_text(encoding="utf-8-sig"))
        yield _matched(address="second")

    assert write_results(probing_results(), path=target, headers=HEADERS) == 2
    assert seen_midway and "first" in seen_midway[0]
    assert "second" not in seen_midway[0]


# ---- XLSX ----

def test_xlsx_round_trip(tmp_path):
    target = tmp_path / "out.xlsx"
    count = write_results(
        [_matched(), _matched(address="233 S Wacker Dr")],
        path=target,
        headers=HEADERS,
        layer_attributes=ATTRIBUTES,
    )
    assert count == 2

    rows = _read_xlsx(target)
    assert tuple(rows[0]) == HEADERS + result_columns(ATTRIBUTES)
    assert rows[1][:3] == ["121 N La Salle St", "manual-7", "C-001"]
    assert rows[2][0] == "233 S Wacker Dr"
    header = list(rows[0])
    assert rows[1][header.index("pip_dist_num")] == "17"
    assert rows[1][header.index("pip_status")] == STATUS_MATCHED
    # Written as text (plan R14) so leading zeros survive the round trip.
    # openpyxl reads an empty cell back as None; every populated one is a str.
    assert all(cell is None or isinstance(cell, str) for cell in rows[1])
    assert rows[1][header.index("pip_score")] == "100.0"


def test_xlsx_preserves_leading_zeros(tmp_path):
    target = tmp_path / "zips.xlsx"
    write_results(
        [
            RowResult(
                row={"address": "1 Main St", "score": "007", "case_id": "0060602"},
                status=STATUS_MATCHED,
            )
        ],
        path=target,
        headers=HEADERS,
    )
    rows = _read_xlsx(target)
    assert rows[1][1] == "007"
    assert rows[1][2] == "0060602"


def test_xlsx_escapes_formula_injection(tmp_path):
    target = tmp_path / "injection.xlsx"
    write_results(
        [
            RowResult(
                row={"address": INJECTION, "score": "", "case_id": ""},
                status=STATUS_MATCHED,
                matched_address=INJECTION,
            )
        ],
        path=target,
        headers=HEADERS,
    )
    rows = _read_xlsx(target)
    header = list(rows[0])
    assert rows[1][header.index("address")] == "'" + INJECTION
    assert rows[1][header.index("pip_matched_address")] == "'" + INJECTION


def test_xlsx_iterator_failure_still_saves_consumed_rows(tmp_path):
    # An XLSX cannot be left half-written the way a CSV can (the zip container is
    # finalized on save), so the writer saves what it consumed before re-raising.
    target = tmp_path / "partial.xlsx"

    def failing_results():
        yield _matched(address="first")
        raise RuntimeError("provider died mid-run")

    with pytest.raises(RuntimeError, match="provider died mid-run"):
        write_results(failing_results(), path=target, headers=HEADERS)

    rows = _read_xlsx(target)
    assert len(rows) == 2
    assert rows[1][0] == "first"


def test_missing_openpyxl_raises_batch_error_with_install_hint(tmp_path, monkeypatch):
    # Simulate the locked-down box that could not install the [batch] extra: the
    # failure must name the fix, not surface a bare ImportError.
    import builtins

    real_import = builtins.__import__

    def blocked_import(name, *args, **kwargs):
        if name == "openpyxl":
            raise ImportError("No module named 'openpyxl'")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", blocked_import)
    with pytest.raises(BatchError, match=r"optional 'batch' extra"):
        write_results([_matched()], path=tmp_path / "out.xlsx", headers=HEADERS)


def test_unwritable_destination_raises_batch_error_not_oserror(tmp_path):
    # A missing parent directory is an operator typo, not a bug. The module's
    # caller contract is BatchError (the CLI catches it and exits 2 without a
    # traceback), so no OSError may escape.
    target = tmp_path / "no-such-dir" / "out.csv"

    with pytest.raises(BatchError) as raised:
        write_results([_matched()], path=target, headers=HEADERS)

    assert "could not open" in str(raised.value)
    assert "out.csv" in str(raised.value)
    # And the directory was NOT conjured up: this feature writes exactly the one
    # file it was told to write and nothing else (SPEC §9).
    assert not target.parent.exists()


def test_unwritable_xlsx_destination_fails_before_any_row_is_consumed(tmp_path):
    target = tmp_path / "no-such-dir" / "out.xlsx"
    consumed = []

    def counting_results():
        for result in [_matched()]:
            consumed.append(result)
            yield result

    with pytest.raises(BatchError) as raised:
        write_results(counting_results(), path=target, headers=HEADERS)

    assert "could not open" in str(raised.value)
    assert not target.parent.exists()
    # openpyxl only touches the path at save(), i.e. at the very end. Without the
    # up-front claim on the destination, a half-hour run would do all its work
    # and then throw every row away.
    assert consumed == []


def test_an_xlsx_run_leaves_no_openpyxl_temp_file_behind(tmp_path):
    # openpyxl stages worksheet XML — the operator's rows, addresses included —
    # in $TMPDIR/openpyxl.* while it builds the workbook, and removes it on
    # save(). SPEC §9 forbids persisting queried addresses, so this pins the
    # cleanup we depend on: nothing may survive a completed run.
    import glob
    import os
    import tempfile

    pattern = os.path.join(tempfile.gettempdir(), "openpyxl.*")
    before = set(glob.glob(pattern))

    write_results(
        [_matched()],
        path=tmp_path / "out.xlsx",
        headers=HEADERS,
        layer_attributes=ATTRIBUTES,
    )

    assert set(glob.glob(pattern)) - before == set()


def test_a_csv_run_writes_exactly_one_file(tmp_path):
    # SPEC §9: the file the operator named is the only artifact of a run. No
    # sibling temp file, no lock file, no cache.
    target = tmp_path / "out.csv"

    write_results(
        [_matched(), _matched()],
        path=target,
        headers=HEADERS,
        layer_attributes=ATTRIBUTES,
    )

    assert [entry.name for entry in tmp_path.iterdir()] == ["out.csv"]
