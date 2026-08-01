"""F7-T1 — tests for the batch source readers.

Fixtures are authored in-test: a CSV written byte-for-byte with the utf-8-sig BOM
Excel emits, a real .xlsx built with openpyxl, and every Google Sheets fetch
mocked with respx. Nothing here touches the network, so the suite passes offline.

The recurring assertion is that a leading zero survives: a ZIP of "07001" read
back as 7001 is the defining data-loss bug of batch geocoding, and it is silent.
"""
import time
import zipfile
from pathlib import Path

import httpx
import pytest
import respx
from openpyxl import Workbook

from app.batch import BatchError, ColumnMapping, Source
from app.batch.sources import (
    parse_google_sheets_url,
    read_source,
    redact_source_spec,
    sheets_export_url,
    validate_column_mapping,
)

CSV_BODY = (
    "Address,ZIP,Case Number\r\n"
    "121 N La Salle St,07001,0042\r\n"
    "233 S Wacker Dr,60606,7\r\n"
)

SHEET_URL = "https://docs.google.com/spreadsheets/d/ABC123_doc-id/edit#gid=1874"
EXPORT_URL = (
    "https://docs.google.com/spreadsheets/d/ABC123_doc-id/export"
    "?format=csv&gid=1874"
)

SIGN_IN_HTML = (
    "<!DOCTYPE html><html><head><title>Sign in - Google Accounts</title></head>"
    "<body>Use your Google Account</body></html>"
)


def _write_csv(tmp_path: Path, body: str = CSV_BODY, *, bom: bool = True) -> Path:
    """Write a CSV exactly as Excel's 'CSV UTF-8' export does: BOM, CRLF."""
    path = tmp_path / "caseload.csv"
    prefix = "﻿" if bom else ""
    path.write_bytes((prefix + body).encode("utf-8"))
    return path


def _write_xlsx(tmp_path: Path, rows: list[list] | None = None) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    for record in rows or [
        ["Address", "ZIP", "Case Number"],
        ["121 N La Salle St", "07001", "0042"],
        ["233 S Wacker Dr", 60606, 7],
    ]:
        worksheet.append(record)
    path = tmp_path / "caseload.xlsx"
    workbook.save(path)
    return path


# ---- CSV ---------------------------------------------------------------


def test_csv_headers_and_rows(tmp_path):
    source = read_source(str(_write_csv(tmp_path)))
    assert source.headers == ("Address", "ZIP", "Case Number")
    assert source.name == "caseload.csv"
    rows = list(source.rows)
    assert len(rows) == 2
    assert rows[0] == {
        "Address": "121 N La Salle St",
        "ZIP": "07001",
        "Case Number": "0042",
    }


def test_csv_bom_is_stripped_from_first_header(tmp_path):
    # Without encoding="utf-8-sig" the first column is named "﻿Address" and
    # every column mapping against it fails for no visible reason.
    source = read_source(str(_write_csv(tmp_path)))
    assert source.headers[0] == "Address"
    assert "﻿" not in "".join(source.headers)


def test_csv_leading_zeros_survive(tmp_path):
    row = next(read_source(str(_write_csv(tmp_path))).rows)
    assert row["ZIP"] == "07001"
    assert row["Case Number"] == "0042"
    assert all(isinstance(value, str) for value in row.values())


def test_csv_without_bom_reads_identically(tmp_path):
    source = read_source(str(_write_csv(tmp_path, bom=False)))
    assert source.headers == ("Address", "ZIP", "Case Number")


def test_csv_short_row_is_padded_not_dropped(tmp_path):
    body = "Address,ZIP\r\n121 N La Salle St\r\n"
    rows = list(read_source(str(_write_csv(tmp_path, body))).rows)
    assert rows == [{"Address": "121 N La Salle St", "ZIP": ""}]


def test_csv_long_row_does_not_abort_the_run(tmp_path):
    body = "Address,ZIP\r\n121 N La Salle St,60602,extra\r\n233 S Wacker Dr,60606\r\n"
    rows = list(read_source(str(_write_csv(tmp_path, body))).rows)
    assert len(rows) == 2
    assert rows[0] == {"Address": "121 N La Salle St", "ZIP": "60602"}


# The defect these tests exist for: three data rows in, two out. Python's CSV
# reader treats the unclosed quote on the A-2 line as a quoted field running to
# EOF, so the whole A-3 line is absorbed into A-2's `notes` and A-3 is gone. The
# runaway field never reaches csv.field_size_limit, so nothing is raised.
UNCLOSED_QUOTE_BODY = (
    "case,notes,lat,lon\n"
    "A-1,fine,41.88,-87.63\n"
    'A-2,"he said,41.87,-87.62\n'
    "A-3,also fine,41.92,-87.65\n"
)

# The same shape, but legitimate: a quoted notes cell that really does contain a
# newline. Indistinguishable from the above to any CSV parser, which is why the
# reader warns rather than repairs.
LEGITIMATE_MULTILINE_BODY = (
    "case,notes,lat,lon\n"
    "B-1,fine,41.88,-87.63\n"
    'B-2,"line one\nline two",41.87,-87.62\n'
    "B-3,also fine,41.92,-87.65\n"
)


def test_csv_unclosed_quote_warns_instead_of_losing_rows_silently(tmp_path):
    warnings: list[str] = []
    source = read_source(
        str(_write_csv(tmp_path, UNCLOSED_QUOTE_BODY)), on_warning=warnings.append
    )
    rows = list(source.rows)

    # The row loss is real and unfixable at this layer — but it is no longer silent.
    assert len(rows) == 2
    assert len(warnings) == 1
    assert "line 3" in warnings[0]
    assert "unclosed" in warnings[0]
    assert "caseload.csv" in warnings[0]


def test_csv_multiline_warning_never_quotes_the_cell(tmp_path):
    """SPEC §9: the swallowed cell is very often the address column."""
    warnings: list[str] = []
    list(
        read_source(
            str(_write_csv(tmp_path, UNCLOSED_QUOTE_BODY)), on_warning=warnings.append
        ).rows
    )
    message = warnings[0]
    for cell_content in ("he said", "A-2", "A-3", "41.87", "-87.62", "also fine"):
        assert cell_content not in message


def test_csv_legitimate_multiline_cell_still_parses_whole(tmp_path):
    """A genuine multi-line cell is warned about but never dropped or mangled."""
    warnings: list[str] = []
    source = read_source(
        str(_write_csv(tmp_path, LEGITIMATE_MULTILINE_BODY)),
        on_warning=warnings.append,
    )
    rows = list(source.rows)

    assert [row["case"] for row in rows] == ["B-1", "B-2", "B-3"]
    assert rows[1]["notes"] == "line one\nline two"
    assert rows[1]["lon"] == "-87.62"
    assert len(warnings) == 1
    assert "line 3" in warnings[0]


def test_csv_warns_once_per_affected_record(tmp_path):
    body = (
        "case,notes\n"
        'A,"one\ntwo"\n'
        "B,plain\n"
        'C,"three\nfour\nfive"\n'
    )
    warnings: list[str] = []
    list(read_source(str(_write_csv(tmp_path, body)), on_warning=warnings.append).rows)
    # Record A occupies lines 2-3, B line 4, C lines 5-7.
    assert len(warnings) == 2
    assert "line 2" in warnings[0]
    assert "line 5" in warnings[1]


def test_csv_without_on_warning_is_unchanged(tmp_path):
    """The callback is optional; omitting it must not alter parsing."""
    rows = list(read_source(str(_write_csv(tmp_path, UNCLOSED_QUOTE_BODY))).rows)
    assert len(rows) == 2


def test_csv_clean_file_produces_no_warning(tmp_path):
    warnings: list[str] = []
    list(read_source(str(_write_csv(tmp_path)), on_warning=warnings.append).rows)
    assert warnings == []


def test_csv_rows_are_lazy(tmp_path):
    source = read_source(str(_write_csv(tmp_path)))
    assert not isinstance(source.rows, list)
    assert isinstance(next(source.rows), dict)


def test_missing_file_raises_batch_error(tmp_path):
    with pytest.raises(BatchError, match="source file not found"):
        read_source(str(tmp_path / "nope.csv"))


def test_empty_file_raises_batch_error(tmp_path):
    path = tmp_path / "empty.csv"
    path.write_text("")
    with pytest.raises(BatchError, match="empty"):
        read_source(str(path))


def test_blank_header_name_raises_batch_error(tmp_path):
    with pytest.raises(BatchError, match="blank column name"):
        read_source(str(_write_csv(tmp_path, "Address,,ZIP\r\na,b,c\r\n")))


def test_duplicate_header_raises_batch_error(tmp_path):
    # Two columns of the same name would silently collapse in a row dict.
    with pytest.raises(BatchError, match="duplicate column name"):
        read_source(str(_write_csv(tmp_path, "Address,ZIP,Address\r\na,b,c\r\n")))


def test_header_whitespace_is_trimmed(tmp_path):
    source = read_source(str(_write_csv(tmp_path, "Address , ZIP\r\na,b\r\n")))
    assert source.headers == ("Address", "ZIP")


def test_non_utf8_file_raises_batch_error(tmp_path):
    path = tmp_path / "latin1.csv"
    path.write_bytes("Address,City\r\n1 Rue Ren\xe9,Montr\xe9al\r\n".encode("latin-1"))
    with pytest.raises(BatchError, match="not valid UTF-8"):
        list(read_source(str(path)).rows)


# ---- XLSX --------------------------------------------------------------


def test_xlsx_headers_and_rows(tmp_path):
    source = read_source(str(_write_xlsx(tmp_path)))
    assert source.headers == ("Address", "ZIP", "Case Number")
    assert source.name == "caseload.xlsx"
    rows = list(source.rows)
    assert len(rows) == 2
    assert rows[0]["ZIP"] == "07001"


def test_xlsx_text_cell_keeps_leading_zero(tmp_path):
    rows = list(read_source(str(_write_xlsx(tmp_path))).rows)
    assert rows[0]["Case Number"] == "0042"


def test_xlsx_numeric_cell_has_no_float_tail(tmp_path):
    # openpyxl reports a numeric ZIP as 60606.0; "60606.0" geocodes to nothing.
    rows = list(read_source(str(_write_xlsx(tmp_path))).rows)
    assert rows[1]["ZIP"] == "60606"
    assert rows[1]["Case Number"] == "7"
    assert all(isinstance(value, str) for value in rows[1].values())


def test_xlsx_empty_cell_becomes_empty_string(tmp_path):
    path = _write_xlsx(
        tmp_path,
        [["Address", "ZIP"], ["121 N La Salle St", None]],
    )
    assert list(read_source(str(path)).rows) == [
        {"Address": "121 N La Salle St", "ZIP": ""}
    ]


def test_xlsx_trailing_blank_rows_are_skipped(tmp_path):
    path = _write_xlsx(
        tmp_path,
        [["Address", "ZIP"], ["121 N La Salle St", "60602"], [None, None], [None, None]],
    )
    assert len(list(read_source(str(path)).rows)) == 1


def test_xlsx_with_no_header_row_raises_batch_error(tmp_path):
    workbook = Workbook()
    path = tmp_path / "blank.xlsx"
    workbook.save(path)
    with pytest.raises(BatchError, match="empty|blank"):
        read_source(str(path))


def test_corrupt_xlsx_raises_batch_error(tmp_path):
    path = tmp_path / "corrupt.xlsx"
    path.write_bytes(b"this is definitely not a zip archive")
    with pytest.raises(BatchError, match="could not read workbook"):
        read_source(str(path))


def test_missing_openpyxl_tells_operator_how_to_install(tmp_path, monkeypatch):
    # Simulate an install without the optional "batch" extra.
    import builtins

    path = _write_xlsx(tmp_path)
    real_import = builtins.__import__

    def fake_import(name, *args, **kwargs):
        if name == "openpyxl":
            raise ImportError("No module named 'openpyxl'")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", fake_import)
    with pytest.raises(BatchError) as error:
        read_source(str(path))
    assert 'pip install -e ".[batch]"' in str(error.value)


# ---- Google Sheets URL parsing -----------------------------------------


def test_gid_in_fragment():
    assert parse_google_sheets_url(SHEET_URL) == ("ABC123_doc-id", "1874")


def test_gid_in_query():
    url = "https://docs.google.com/spreadsheets/d/DOC9/export?format=csv&gid=55"
    assert parse_google_sheets_url(url) == ("DOC9", "55")


def test_gid_absent_defaults_to_first_tab():
    url = "https://docs.google.com/spreadsheets/d/DOC9/edit"
    assert parse_google_sheets_url(url) == ("DOC9", "0")


def test_explicit_sheet_gid_wins_over_url():
    assert parse_google_sheets_url(SHEET_URL, sheet_gid="7") == ("ABC123_doc-id", "7")


@pytest.mark.parametrize(
    "url",
    [
        "https://example.com/spreadsheets/d/DOC9/edit",   # wrong host
        "https://docs.google.com/document/d/DOC9/edit",   # a Doc, not a Sheet
        "https://docs.google.com/spreadsheets/",          # no document id
        "https://docs.google.com/",
    ],
)
def test_non_sheets_url_raises_batch_error(url):
    with pytest.raises(BatchError, match="not a Google Sheets URL"):
        read_source(url)


# ---- Google Sheets fetch -----------------------------------------------


@respx.mock
def test_sheet_happy_path_parses_csv():
    respx.get(EXPORT_URL).mock(
        return_value=httpx.Response(
            200, text=CSV_BODY, headers={"content-type": "text/csv; charset=utf-8"}
        )
    )
    source = read_source(SHEET_URL)
    assert source.headers == ("Address", "ZIP", "Case Number")
    rows = list(source.rows)
    assert len(rows) == 2
    assert rows[0]["ZIP"] == "07001"  # leading zero survives the round trip


@respx.mock
def test_sheet_request_uses_the_csv_export_endpoint():
    route = respx.get(EXPORT_URL).mock(
        return_value=httpx.Response(200, text=CSV_BODY, headers={"content-type": "text/csv"})
    )
    list(read_source(SHEET_URL).rows)
    url = str(route.calls.last.request.url)
    assert "/export" in url and "format=csv" in url and "gid=1874" in url


@respx.mock
def test_sheet_gid_argument_overrides_url_gid():
    override_url = (
        "https://docs.google.com/spreadsheets/d/ABC123_doc-id/export?format=csv&gid=9"
    )
    respx.get(override_url).mock(
        return_value=httpx.Response(200, text=CSV_BODY, headers={"content-type": "text/csv"})
    )
    source = read_source(SHEET_URL, sheet_gid="9")
    assert source.headers[0] == "Address"


@respx.mock
def test_sheet_not_shared_returns_login_html_not_csv():
    # Google answers 200 with a sign-in PAGE, which is valid one-column CSV.
    # Undetected, that is thousands of nonsense rows and thousands of geocodes.
    respx.get(EXPORT_URL).mock(
        return_value=httpx.Response(
            200, text=SIGN_IN_HTML, headers={"content-type": "text/html; charset=utf-8"}
        )
    )
    with pytest.raises(BatchError) as error:
        read_source(SHEET_URL)
    message = str(error.value)
    assert "not link-shared" in message
    assert "Anyone with the link" in message
    assert "Viewer" in message


@respx.mock
def test_sheet_not_shared_detected_without_html_content_type():
    # Same page served with a generic content type: the body still opens as markup.
    respx.get(EXPORT_URL).mock(
        return_value=httpx.Response(
            200, text=SIGN_IN_HTML, headers={"content-type": "application/octet-stream"}
        )
    )
    with pytest.raises(BatchError, match="not link-shared"):
        read_source(SHEET_URL)


@respx.mock
def test_sheet_server_error_is_distinct_from_not_shared():
    respx.get(EXPORT_URL).mock(return_value=httpx.Response(500, text="<html>oops</html>"))
    with pytest.raises(BatchError) as error:
        read_source(SHEET_URL)
    message = str(error.value)
    assert "HTTP 500" in message
    assert "not link-shared" not in message


@respx.mock
def test_sheet_network_failure_raises_batch_error_not_httpx():
    respx.get(EXPORT_URL).mock(side_effect=httpx.ConnectError("no route to host"))
    with pytest.raises(BatchError) as error:
        read_source(SHEET_URL)
    assert "could not reach Google Sheets" in str(error.value)


@respx.mock
def test_sheet_timeout_raises_batch_error_not_httpx():
    respx.get(EXPORT_URL).mock(side_effect=httpx.ReadTimeout("timed out"))
    with pytest.raises(BatchError, match="could not reach"):
        read_source(SHEET_URL)


# ---- regression: interior blank rows must not vanish (S1) ---------------

BLANK_GAP_ROWS = [
    ["Address", "ZIP"],
    ["121 N La Salle St", "60602"],
    ["233 S Wacker Dr", "60606"],
    [None, None],
    ["30 N La Salle St", "60602"],
    ["1 N State St", "60602"],
]
BLANK_GAP_CSV = (
    "Address,ZIP\r\n"
    "121 N La Salle St,60602\r\n"
    "233 S Wacker Dr,60606\r\n"
    "\r\n"
    "30 N La Salle St,60602\r\n"
    "1 N State St,60602\r\n"
)


def test_xlsx_interior_blank_row_is_kept_not_skipped(tmp_path):
    # Dropping it shifts every later row up by one, so an operator pasting the
    # pip_* columns back beside their own sheet gets the wrong district on every
    # row after the gap — silently, with exit code 0.
    rows = list(read_source(str(_write_xlsx(tmp_path, BLANK_GAP_ROWS))).rows)
    assert len(rows) == 5
    assert rows[2] == {"Address": "", "ZIP": ""}
    assert rows[3] == {"Address": "30 N La Salle St", "ZIP": "60602"}


def test_xlsx_blank_rows_after_the_last_data_row_are_still_dropped(tmp_path):
    # "Trailing" is about position: a blank row is padding only if nothing
    # follows it.
    path = _write_xlsx(
        tmp_path,
        [
            ["Address", "ZIP"],
            ["121 N La Salle St", "60602"],
            [None, None],
            ["30 N La Salle St", "60602"],
            [None, None],
            [None, None],
        ],
    )
    rows = list(read_source(str(path)).rows)
    assert len(rows) == 3
    assert rows[-1] == {"Address": "30 N La Salle St", "ZIP": "60602"}


def test_csv_and_xlsx_agree_on_row_count_for_identical_content(tmp_path):
    csv_rows = list(read_source(str(_write_csv(tmp_path, BLANK_GAP_CSV))).rows)
    xlsx_rows = list(read_source(str(_write_xlsx(tmp_path, BLANK_GAP_ROWS))).rows)
    assert csv_rows == xlsx_rows


# ---- regression: a malformed CSV is a BatchError, not a traceback (S2) ---


def _write_unclosed_quote_csv(tmp_path: Path, *, on_header: bool = False) -> Path:
    """A CSV with one unclosed double quote — a routine export artifact.

    The reader then runs past every delimiter looking for the closing quote and
    raises `csv.Error("field larger than field limit")`, which needs the runaway
    field to exceed csv's 128 KiB limit; hence the padding.
    """
    runaway = "x" * 200_000
    if on_header:
        body = f'"Address,ZIP\r\n{runaway}\r\n'
    else:
        body = f'Address,ZIP\r\n"121 N La Salle St,60602\r\n{runaway}\r\n'
    return _write_csv(tmp_path, body, bom=False)


def test_malformed_csv_raises_batch_error_mid_stream(tmp_path):
    # Uncaught, csv.Error escapes the generator, the runner, the writer and every
    # except clause in the CLI: a traceback plus exit 1, which the CLI documents
    # as "finished, some rows unmatched".
    source = read_source(str(_write_unclosed_quote_csv(tmp_path)))
    with pytest.raises(BatchError) as error:
        list(source.rows)
    message = str(error.value)
    assert "could not parse" in message
    assert "caseload.csv" in message


def test_malformed_csv_error_names_the_row_but_never_its_contents(tmp_path):
    source = read_source(str(_write_unclosed_quote_csv(tmp_path)))
    with pytest.raises(BatchError) as error:
        list(source.rows)
    message = str(error.value)
    assert "line 2" in message  # header is line 1
    assert "121 N La Salle St" not in message  # SPEC §9: no address in a message


def test_malformed_csv_header_row_raises_batch_error(tmp_path):
    with pytest.raises(BatchError, match="could not parse"):
        read_source(str(_write_unclosed_quote_csv(tmp_path, on_header=True)))


# ---- regression: no credential in any message (S3) ----------------------

PRESIGNED_URL = (
    "https://files.example.com/export/caseload.csv"
    "?token=SECRET-BEARER-abc123&X-Amz-Signature=deadbeef#frag"
)


def test_non_sheets_url_error_does_not_echo_the_query_string():
    # An operator pasting a presigned S3/Dropbox/SharePoint link must not find
    # its token in terminal scrollback or a CI log.
    with pytest.raises(BatchError) as error:
        read_source(PRESIGNED_URL)
    message = str(error.value)
    assert "SECRET-BEARER-abc123" not in message
    assert "X-Amz-Signature" not in message
    assert "deadbeef" not in message
    assert "files.example.com/export/caseload.csv" in message


def test_missing_file_error_does_not_echo_a_query_string(tmp_path):
    # A mistyped scheme routes a URL down the file-path branch; the message
    # must be just as safe there.
    with pytest.raises(BatchError) as error:
        read_source("htps://files.example.com/x.csv?token=SECRET-BEARER-abc123")
    assert "SECRET-BEARER-abc123" not in str(error.value)


def test_redact_source_spec_drops_query_fragment_and_userinfo():
    assert redact_source_spec(PRESIGNED_URL) == (
        "https://files.example.com/export/caseload.csv"
    )
    assert (
        redact_source_spec("https://user:pw@example.com/a.csv")
        == "https://example.com/a.csv"
    )
    assert redact_source_spec("/data/caseload.csv") == "/data/caseload.csv"


# ---- regression: "Publish to the web" URLs (S4) -------------------------

PUBLISHED_URL = (
    "https://docs.google.com/spreadsheets/d/e/2PACX-1vABCdef_123/pubhtml#gid=77"
)
PUBLISHED_EXPORT_URL = (
    "https://docs.google.com/spreadsheets/d/e/2PACX-1vABCdef_123/pub"
    "?output=csv&gid=77"
)


def test_published_url_does_not_parse_the_document_id_as_e():
    # "/d/e/<id>/pub" is what File > Share > Publish to web hands out; reading
    # it as the private form yields the literal document id "e" and a 404 that
    # blames a URL that was correct.
    document_id, gid = parse_google_sheets_url(PUBLISHED_URL)
    assert document_id != "e"
    assert "2PACX-1vABCdef_123" in document_id
    assert gid == "77"


def test_published_url_exports_through_the_pub_endpoint():
    document_id, gid = parse_google_sheets_url(PUBLISHED_URL)
    assert sheets_export_url(document_id, gid) == PUBLISHED_EXPORT_URL


def test_private_url_still_exports_through_the_export_endpoint():
    document_id, gid = parse_google_sheets_url(SHEET_URL)
    assert document_id == "ABC123_doc-id"
    assert sheets_export_url(document_id, gid) == EXPORT_URL


@respx.mock
def test_published_sheet_is_fetched_and_parsed():
    respx.get(PUBLISHED_EXPORT_URL).mock(
        return_value=httpx.Response(
            200, text=CSV_BODY, headers={"content-type": "text/csv"}
        )
    )
    source = read_source(PUBLISHED_URL)
    assert source.headers == ("Address", "ZIP", "Case Number")
    assert len(list(source.rows)) == 2


# ---- regression: a workbook that expands absurdly is refused (S5) -------


def _write_zip_bomb_xlsx(tmp_path: Path) -> Path:
    """A structurally valid zip whose shared-string table expands ~1000x.

    openpyxl reads xl/sharedStrings.xml fully into memory before yielding a
    single row, so --max-rows cannot intervene: the memory is spent inside
    load_workbook.
    """
    path = tmp_path / "bomb.xlsx"
    payload = ("<si><t>" + "A" * 1000 + "</t></si>") * 8000  # ~8 MB, ~8 KB packed
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "<Types/>")
        archive.writestr("xl/sharedStrings.xml", payload)
    return path


def test_zip_bomb_workbook_is_refused_before_openpyxl_opens_it(tmp_path):
    path = _write_zip_bomb_xlsx(tmp_path)
    started = time.monotonic()
    with pytest.raises(BatchError) as error:
        read_source(str(path))
    elapsed = time.monotonic() - started

    message = str(error.value)
    assert "refusing to open workbook" in message
    assert "bomb.xlsx" in message
    assert ".csv" in message  # tells the operator what to do instead
    # The guard reads the zip directory only, so it cannot be slow.
    assert elapsed < 2.0


def test_zip_bomb_guard_leaves_a_real_workbook_alone(tmp_path):
    # The bound must never fire on an ordinary spreadsheet.
    assert len(list(read_source(str(_write_xlsx(tmp_path))).rows)) == 2


# ---- column-mapping validation -----------------------------------------


def _source(*headers: str) -> Source:
    return Source(headers=tuple(headers), rows=iter([]), name="caseload.csv")


def test_validate_accepts_a_present_address_column():
    validate_column_mapping(_source("Address", "ZIP"), ColumnMapping(address="Address"))


def test_validate_accepts_present_lat_lon_columns():
    source = _source("lat", "lon", "ZIP")
    validate_column_mapping(source, ColumnMapping(lat="lat", lon="lon"))


def test_validate_names_the_missing_column_and_lists_the_real_ones():
    # The operator mistyped "Address" — show them the actual spelling.
    with pytest.raises(BatchError) as error:
        validate_column_mapping(
            _source("Address", "ZIP", "Case Number"), ColumnMapping(address="Adress")
        )
    message = str(error.value)
    assert "'Adress'" in message
    assert "'Address'" in message
    assert "'ZIP'" in message
    assert "'Case Number'" in message
    assert "caseload.csv" in message


def test_validate_reports_every_missing_coordinate_column():
    with pytest.raises(BatchError) as error:
        validate_column_mapping(
            _source("Address"), ColumnMapping(lat="Latitude", lon="Longitude")
        )
    message = str(error.value)
    assert "'Latitude'" in message
    assert "'Longitude'" in message


def test_validate_is_case_sensitive_and_never_guesses():
    # No fuzzy matching (plan D18): "address" is not "Address".
    with pytest.raises(BatchError):
        validate_column_mapping(_source("Address"), ColumnMapping(address="address"))


def test_validate_against_a_real_csv_source(tmp_path):
    source = read_source(str(_write_csv(tmp_path)))
    validate_column_mapping(source, ColumnMapping(address="Address"))
    with pytest.raises(BatchError, match="Columns in this source"):
        validate_column_mapping(source, ColumnMapping(address="street_address"))
