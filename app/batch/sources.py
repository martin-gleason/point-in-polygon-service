"""F7-T1 — reading a batch source: a CSV, an XLSX, or a published Google Sheet.

`read_source(spec)` is the one entry point. `spec` is whatever the operator
typed: a path on disk, or a Google Sheets URL pasted straight out of the browser
address bar. This module dispatches on it, and hands back a `Source` — the
header row in the operator's original order, plus a *lazy* iterator of rows.

Three rules from the package contract shape every backend here:

- **Every cell is a `str`** (plan R14). A ZIP code of "07001" and a house number
  of "0042" must survive intact; the moment a reader coerces a cell to `int` the
  leading zeros are gone and nothing downstream can recover them. No reader in
  this module ever infers a type.
- **Nothing is persisted** (SPEC §9, plan D15). No temp file, no download cache,
  no scratch copy of the operator's spreadsheet. The Google Sheets body is
  parsed from memory; the local backends stream from the operator's own file.
- **No address in a log record or an exception message** (SPEC §9). Errors here
  name the *file*, the *column*, and the *row number* — never a cell's contents.
  The same rule covers the source spec itself: a URL reaches a message only
  through `redact_source_spec`, which drops the query string, the parameters,
  the fragment and any ``user:password@`` userinfo. An operator who pastes a
  presigned link must not find its signature in terminal scrollback or a CI log.

ArcGIS / ArcPy equivalent
    This is the input side of `arcpy.geocoding.geocodeAddresses`: pointing the
    tool at an address table, which in ArcMap/Pro means an OLE DB or Excel
    connection, or `arcpy.conversion.ExcelToTable` to stage the sheet into a
    geodatabase table first. The Esri path materializes an intermediate table on
    disk and lets the Excel driver guess each column's type (the classic cause of
    ZIP codes losing their leading zero). Here there is no intermediate table and
    no type inference: rows stream from the file as text and go straight into the
    run. The Google Sheets backend has no Esri counterpart at all — the nearest
    thing is manually downloading the sheet to .csv first.
"""
from __future__ import annotations

import csv
import io
import re
import zipfile
from collections.abc import Callable, Iterator
from pathlib import Path
from urllib.parse import parse_qs, urlparse, urlunparse

import httpx

from app.batch import BatchError, ColumnMapping, Source

# Workbook suffixes routed to the openpyxl backend; anything else on disk is
# read as delimited text.
XLSX_SUFFIXES = (".xlsx", ".xlsm")

# Google's "publish a sheet as CSV" endpoint. This is the documented export URL,
# not a private API, and it needs no key — consistent with the key-free runtime
# rule. It answers with CSV only when the sheet is link-shared.
SHEETS_EXPORT_TEMPLATE = (
    "https://docs.google.com/spreadsheets/d/{doc_id}/export?format=csv&gid={gid}"
)

# The other shape Google hands out, under File > Share > Publish to web:
# ``/spreadsheets/d/e/<PUBLISHED_ID>/pub...``. It is a different document id in a
# different namespace — the "e" is a literal path segment, not a document id —
# and it exports through /pub, not /export. Read as if it were the private form
# the document id parses out as the bare letter "e" and every fetch 404s while
# blaming a URL that was correct.
SHEETS_PUBLISHED_EXPORT_TEMPLATE = (
    "https://docs.google.com/spreadsheets/d/e/{published_id}/pub?output=csv&gid={gid}"
)
SHEETS_HOST_SUFFIX = "docs.google.com"

# A published id is carried through the parse prefixed with "e/" so one
# (document id, gid) pair can describe either shape; `sheets_export_url` is the
# single place that turns it back into a fetchable URL.
PUBLISHED_ID_PREFIX = "e/"
_SHEETS_PUBLISHED_ID = re.compile(r"/spreadsheets/d/e/([A-Za-z0-9_-]+)")
_SHEETS_DOC_ID = re.compile(r"/spreadsheets/d/([A-Za-z0-9_-]+)")

# Bounds on what a .xlsx may expand to, checked against the zip directory before
# openpyxl opens the file. openpyxl streams the *sheet*, but it reads
# xl/sharedStrings.xml fully into memory first, so `--max-rows` cannot intervene:
# a structurally valid 320 KB workbook whose shared-string table expands to
# 112 MB costs ~445 MB of RSS before the first row is yielded, and a 10 MB one
# lands in the tens of gigabytes.
#
# The numbers are generous against real spreadsheets. A 100,000-row, 20-column
# workbook of addresses unpacks to roughly 60-80 MB of XML in about a dozen
# members, and its overall compression ratio is around 10:1 (XML markup with
# non-repeating text does not compress much better than that). A ratio of 200:1
# is not something a caseload spreadsheet produces; it is what a table of one
# repeated string produces.
MAX_XLSX_UNCOMPRESSED_BYTES = 256 * 1024 * 1024
MAX_XLSX_MEMBER_UNCOMPRESSED_BYTES = 128 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 200

# Below this, a high ratio means nothing: a 40-byte member of all spaces can
# compress 20:1 and be entirely legitimate.
_RATIO_CHECK_MINIMUM_COMPRESSED_BYTES = 4096

# Seconds to wait on the Sheets export. Generous: Google materializes the CSV on
# demand, and a few thousand rows is not instant.
SHEETS_TIMEOUT_SECONDS = 30.0

# What the operator must do in the Sheets UI when the export comes back as a
# sign-in page. Spelled out verbatim because "not shared" is the single most
# common failure of this backend and the fix is not guessable.
SHARE_INSTRUCTIONS = (
    "In Google Sheets: Share > General access > Anyone with the link > Viewer."
)


def read_source(
    spec: str,
    *,
    sheet_gid: str | None = None,
    on_warning: Callable[[str], None] | None = None,
) -> Source:
    """Read a batch source named by `spec`: a file path or a Google Sheets URL.

    `sheet_gid` selects a tab of a Google Sheet and wins over any `gid` in the
    URL; it is ignored for local files. Raises `BatchError` if the source cannot
    be read — a missing file, an unshared sheet, an empty header row.

    `on_warning` receives a message for each condition that is *suspicious but
    not an error* — today, a CSV record whose cell spans several physical lines
    (see `_multiline_cell_message`). It is a callback rather than a `print`
    because this module writes to no stream of its own: the caller owns the
    decision of where a diagnostic goes, and a library that prints cannot be
    embedded behind an HTTP endpoint. Warnings fire lazily, as the rows stream.
    Messages passed to it name the file, the line and the count only — never a
    cell's contents (SPEC §9).

    The returned `Source.rows` is a generator that holds the file (or the parsed
    body) open until it is exhausted or closed, so a large workbook streams
    rather than landing in memory whole.
    """
    if is_url(spec):
        return _read_google_sheet(spec, sheet_gid=sheet_gid, on_warning=on_warning)

    path = Path(spec).expanduser()
    if not path.is_file():
        # Redacted: a URL typed with a broken scheme ("htps://...") lands here,
        # and its query string may be a credential.
        raise BatchError(f"source file not found: {redact_source_spec(str(path))}")
    if path.suffix.lower() in XLSX_SUFFIXES:
        return _read_xlsx(path)
    return _read_csv(path, on_warning=on_warning)


def validate_column_mapping(source: Source, mapping: ColumnMapping) -> None:
    """Check that every column the mapping needs exists in the source headers.

    Raises `BatchError` naming the missing column *and* listing the headers the
    file actually has. We never fuzzy-match or auto-guess a column (plan D18):
    guessing wrong on a few thousand real addresses produces confident,
    plausible, unnoticed garbage. An operator who typed "Adress" instead of
    "Address" gets to see the real spelling and fix it themselves.
    """
    missing = [
        column for column in mapping.required_columns() if column not in source.headers
    ]
    if not missing:
        return

    label = "columns" if len(missing) > 1 else "column"
    missing_names = ", ".join(repr(column) for column in missing)
    present = ", ".join(repr(header) for header in source.headers) or "(none)"
    raise BatchError(
        f"{source.name}: {label} {missing_names} not found. "
        f"Columns in this source: {present}"
    )


def parse_google_sheets_url(url: str, *, sheet_gid: str | None = None) -> tuple[str, str]:
    """Pull (document id, gid) out of a browser Google Sheets URL.

    Accepts the URL as copied from the address bar, e.g.
    ``https://docs.google.com/spreadsheets/d/<DOCID>/edit#gid=123``, and the
    "Publish to web" shape,
    ``https://docs.google.com/spreadsheets/d/e/<PUBLISHED_ID>/pubhtml``. The two
    are different namespaces with different export endpoints, so a published id
    comes back prefixed with ``"e/"`` and `sheets_export_url` builds the right
    URL for each. The tab id may live in the fragment (what the browser writes)
    or the query string (what a shared "export" link writes); absent, it defaults
    to ``"0"``, the first tab. An explicit `sheet_gid` argument overrides both.

    Raises `BatchError` if the URL is not a Google Sheets document URL. The
    message carries only the redacted URL (`redact_source_spec`): the query
    string of a pasted link can hold a bearer token or an object signature, and
    this message is printed to stderr.
    """
    parsed = urlparse(url)
    host = parsed.netloc.split("@")[-1].split(":")[0].lower()
    is_sheets_host = host == SHEETS_HOST_SUFFIX or host.endswith(
        "." + SHEETS_HOST_SUFFIX
    )
    published_match = _SHEETS_PUBLISHED_ID.search(parsed.path)
    match = published_match or _SHEETS_DOC_ID.search(parsed.path)
    if not is_sheets_host or match is None:
        raise BatchError(
            f"not a Google Sheets URL: {redact_source_spec(url)!r}. Expected "
            "something like "
            "https://docs.google.com/spreadsheets/d/<DOCID>/edit#gid=0"
        )

    document_id = match.group(1)
    if published_match is not None:
        document_id = PUBLISHED_ID_PREFIX + document_id

    if sheet_gid is not None:
        return document_id, str(sheet_gid)
    gid = _first_query_value(parsed.fragment, "gid") or _first_query_value(
        parsed.query, "gid"
    )
    return document_id, gid or "0"


def sheets_export_url(document_id: str, gid: str) -> str:
    """The key-free CSV export URL for a document id as `parse_google_sheets_url`
    returns it — the ``/export`` endpoint for a normal document, the ``/pub``
    endpoint for a published one (the ``"e/"`` prefix is the discriminator)."""
    if document_id.startswith(PUBLISHED_ID_PREFIX):
        return SHEETS_PUBLISHED_EXPORT_TEMPLATE.format(
            published_id=document_id[len(PUBLISHED_ID_PREFIX):], gid=gid
        )
    return SHEETS_EXPORT_TEMPLATE.format(doc_id=document_id, gid=gid)


def redact_source_spec(spec: str) -> str:
    """A source spec reduced to what is safe to print: no query string, no
    parameters, no fragment, no ``user:password@`` userinfo.

    Presigned S3, Dropbox and SharePoint links carry their whole authorization in
    the query string, and an operator who pastes one has no reason to expect it
    in terminal scrollback or a CI log. Non-URL specs (a filesystem path) pass
    through with the same three delimiters cut, because a mistyped scheme routes
    a URL down the file-path branch and its message must be just as safe.
    """
    parsed = urlparse(spec)
    if not parsed.scheme and not parsed.netloc:
        return spec.split("?")[0].split("#")[0]
    return urlunparse(
        parsed._replace(
            netloc=parsed.netloc.split("@")[-1], params="", query="", fragment=""
        )
    )


# ---- backends ----------------------------------------------------------


def _read_csv(
    path: Path, *, on_warning: Callable[[str], None] | None = None
) -> Source:
    """Stream a delimited text file, every cell as a `str`.

    Opened as ``utf-8-sig`` so the byte-order mark Excel writes at the head of a
    "CSV UTF-8" export is consumed rather than becoming part of the first column
    name — otherwise the header reads as ``"\ufeffAddress"`` and every column
    mapping against it mysteriously fails.
    """
    try:
        handle = open(path, newline="", encoding="utf-8-sig")
    except OSError as error:
        raise BatchError(
            f"could not open {redact_source_spec(str(path))}: {error.strerror}"
        ) from error

    try:
        reader = csv.reader(handle)
        headers = _headers_from(next(reader, None), source_name=path.name)
    except UnicodeDecodeError as error:
        handle.close()
        raise BatchError(
            f"{path.name} is not valid UTF-8 text; re-save it as CSV UTF-8 "
            f"(or use the .xlsx directly)"
        ) from error
    except csv.Error as error:
        handle.close()
        raise BatchError(_malformed_csv_message(path.name, row_number=1)) from error
    except BatchError:
        handle.close()
        raise

    return Source(
        headers=headers,
        rows=_stream_csv_rows(handle, reader, headers, path.name, on_warning),
        name=path.name,
    )


def _stream_csv_rows(
    handle: io.TextIOBase,
    reader,
    headers: tuple[str, ...],
    source_name: str,
    on_warning: Callable[[str], None] | None = None,
) -> Iterator[dict[str, str]]:
    """Yield each remaining CSV record as a header→text dict, then close the file.

    The `finally` also runs when the consumer abandons the generator (a `break`
    in the caller, or an error mid-run), so the operator's file is never left
    open by a partial batch.

    Both decode failures and *parse* failures become `BatchError`. A single
    unclosed double quote — a routine export artifact — makes `csv.reader`
    swallow the rest of the file and raise `csv.Error("field larger than field
    limit")` mid-stream. Uncaught, that escapes the generator, sails past every
    `except` clause in the CLI and prints a traceback with exit code 1, which the
    CLI documents as "finished, some rows unmatched": a 4,000-row run that wrote
    two rows looks like an ordinary partial result.

    Below `csv.field_size_limit` — the ordinary case for a small file — the same
    unclosed quote raises nothing at all: the runaway field simply swallows the
    following lines and those rows vanish from the stream. That is why
    `_records_warning_on_multiline_cells` sits in front of this loop.
    """
    row_number = 1  # the header, already consumed
    try:
        for record in _records_warning_on_multiline_cells(
            reader, source_name, on_warning
        ):
            row_number += 1
            yield _row_from_cells(headers, record)
    except UnicodeDecodeError as error:
        raise BatchError(
            f"{source_name} is not valid UTF-8 text; re-save it as CSV UTF-8"
        ) from error
    except csv.Error as error:
        # row_number + 1 is the record that failed to parse: the loop increments
        # only once a record has been handed back intact.
        raise BatchError(
            _malformed_csv_message(source_name, row_number=row_number + 1)
        ) from error
    finally:
        handle.close()


def _records_warning_on_multiline_cells(
    reader,
    source_name: str,
    on_warning: Callable[[str], None] | None,
) -> Iterator[list[str]]:
    """Yield each record from `reader`, warning once per record that absorbed
    more than one physical line.

    This is the only defence against the worst failure this reader has. A single
    unclosed double quote — a routine artifact of exporting free-text notes — is
    read by `csv.reader` as a quoted field that runs on until it finds its
    partner, so every line up to that partner is eaten into one cell and those
    data rows disappear from the stream. Nothing is raised, nothing is logged;
    the operator sees a smaller row count and no reason for it. Because the
    documented workflow is pasting the appended ``pip_*`` columns back beside the
    original spreadsheet, one vanished row shifts every later district onto the
    wrong case.

    It cannot be fixed by detection, because there is nothing to detect: an
    unclosed quote and a legitimate multi-line cell are *the same thing* to any
    CSV parser — a quoted field containing newlines. Guessing between them would
    either drop legitimate cells or miss real corruption. So the reader neither
    guesses nor repairs. It reports, once per affected record, and lets the
    person who knows whether their spreadsheet uses multi-line cells decide.

    `reader.line_num` is the count of *physical* lines consumed so far, which is
    what makes both numbers available: the difference across one iteration is how
    many lines the record absorbed, and the previous value plus one is the line
    it started on. Warnings are skipped entirely when `on_warning` is None, so
    the scan costs nothing to a caller that does not want it.
    """
    previous_line_number = reader.line_num  # the header row, already consumed
    for record in reader:
        if on_warning is not None and any(
            "\n" in cell or "\r" in cell for cell in record
        ):
            on_warning(
                _multiline_cell_message(
                    source_name,
                    start_line=previous_line_number + 1,
                    line_count=reader.line_num - previous_line_number,
                )
            )
        previous_line_number = reader.line_num
        yield record


def _multiline_cell_message(
    source_name: str, *, start_line: int, line_count: int
) -> str:
    """Describe one multi-line record by position and size only.

    Names the file, the line the record started on and how many lines it ate.
    The cell itself is never quoted: the field that swallowed the rest of the
    file is very often the address column, and SPEC §9 keeps addresses out of
    every message, log record and stream.
    """
    return (
        f"{source_name}: the record starting at line {start_line} has a cell "
        f"spanning {line_count} lines. This is normal ONLY if this source "
        f"genuinely uses multi-line cells. Otherwise an unclosed \" quote has "
        f"merged those {line_count} lines into one record and rows are missing "
        f"from this run. Open the file at line {start_line} and check the "
        f"quoting."
    )


def _malformed_csv_message(source_name: str, *, row_number: int) -> str:
    """Explain a `csv.Error` by position only.

    The offending text is deliberately absent: the cell that broke the parse is
    very often the address column, and SPEC §9 keeps addresses out of every
    message, log record and stream.
    """
    return (
        f"{source_name}: could not parse the row starting at line {row_number}. "
        f"The usual cause is an unclosed \" quote, which makes the reader run to "
        f"the end of the file looking for its partner. Open the file at that row "
        f"and fix the quoting, or re-export it from the spreadsheet."
    )


def _read_xlsx(path: Path) -> Source:
    """Stream the first worksheet of an .xlsx workbook, every cell as a `str`.

    openpyxl is imported here rather than at module scope so the CSV and Sheets
    backends work in an install without the optional extra, and so an operator
    who never touches Excel never pays for the import.

    ``read_only=True`` streams the sheet row by row instead of building the whole
    object graph; ``data_only=True`` reads the cached *value* of a formula cell
    rather than the formula text, which is what a caseload spreadsheet with a
    concatenated address column needs.

    A note on leading zeros: a workbook that already stored "07001" as the number
    7001 lost the zero in Excel, before this program ever saw it. We faithfully
    render what the cell holds and never invent padding.

    The workbook's zip directory is checked (`_reject_oversized_workbook`) before
    openpyxl opens it, because `read_only=True` does not make the *whole* read
    incremental — see that function.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise BatchError(
            "reading .xlsx needs the optional 'batch' extra. Install it with: "
            'pip install -e ".[batch]"  (or use a .csv export instead)'
        ) from error

    _reject_oversized_workbook(path)

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as error:  # openpyxl raises a wide family on a bad file
        raise BatchError(
            f"could not read workbook {path.name}: {type(error).__name__}"
        ) from error

    worksheet = workbook.active
    if worksheet is None:
        workbook.close()
        raise BatchError(f"{path.name} has no worksheets")

    row_iterator = worksheet.iter_rows(values_only=True)
    try:
        headers = _headers_from(
            _as_text_cells(next(row_iterator, None)), source_name=path.name
        )
    except BatchError:
        workbook.close()
        raise

    return Source(
        headers=headers,
        rows=_stream_xlsx_rows(workbook, row_iterator, headers),
        name=path.name,
    )


def _reject_oversized_workbook(path: Path) -> None:
    """Refuse a .xlsx whose zip directory promises an absurd expansion.

    ``read_only=True`` streams the *worksheet*, but openpyxl reads
    ``xl/sharedStrings.xml`` into memory in full before the first row is
    available, so a workbook can exhaust memory inside `load_workbook` — before
    a single row exists for ``--max-rows`` to count. A structurally valid 320 KB
    file with a highly repetitive shared-string table expands to 112 MB of XML
    and costs about 445 MB of RSS; at 10 MB of input it reaches the tens of
    gigabytes, and the resulting `MemoryError` is not caught anywhere in the CLI.
    That is a hostile emailed spreadsheet locally, and an unauthenticated request
    body once this reader sits behind an endpoint.

    So the guard is the zip *directory*, which is metadata: no member is
    decompressed to apply it. Three bounds, all far outside anything a real
    caseload workbook produces — see `MAX_XLSX_UNCOMPRESSED_BYTES`,
    `MAX_XLSX_MEMBER_UNCOMPRESSED_BYTES` and `MAX_XLSX_COMPRESSION_RATIO`.

    A file that is not a zip at all is left alone: `load_workbook` already gives
    that case its own "could not read workbook" message.

    ArcGIS / ArcPy equivalent
        None. `arcpy.conversion.ExcelToTable` hands the file to the Excel driver
        and inherits whatever that driver does with it.
    """
    try:
        with zipfile.ZipFile(path) as archive:
            members = archive.infolist()
    except zipfile.BadZipFile:
        return  # not a workbook at all; load_workbook reports it
    except OSError as error:
        raise BatchError(
            f"could not open {redact_source_spec(str(path))}: {error.strerror}"
        ) from error

    total_uncompressed = sum(member.file_size for member in members)
    total_compressed = sum(member.compress_size for member in members)

    if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
        raise BatchError(_oversized_workbook_message(path.name, total_uncompressed))

    for member in members:
        if member.file_size > MAX_XLSX_MEMBER_UNCOMPRESSED_BYTES:
            raise BatchError(_oversized_workbook_message(path.name, member.file_size))
        if (
            member.compress_size >= _RATIO_CHECK_MINIMUM_COMPRESSED_BYTES
            and member.file_size
            > member.compress_size * MAX_XLSX_COMPRESSION_RATIO
        ):
            raise BatchError(_oversized_workbook_message(path.name, member.file_size))

    if (
        total_compressed >= _RATIO_CHECK_MINIMUM_COMPRESSED_BYTES
        and total_uncompressed > total_compressed * MAX_XLSX_COMPRESSION_RATIO
    ):
        raise BatchError(_oversized_workbook_message(path.name, total_uncompressed))


def _oversized_workbook_message(source_name: str, uncompressed_bytes: int) -> str:
    """One message for every bound in `_reject_oversized_workbook`: which file,
    how far it expands, and what to do instead. Names no cell contents."""
    megabytes = uncompressed_bytes / (1024 * 1024)
    return (
        f"refusing to open workbook {source_name}: it expands to about "
        f"{megabytes:,.0f} MB, past the {MAX_XLSX_UNCOMPRESSED_BYTES // (1024 * 1024)} "
        f"MB limit this reader will decompress. Reading it would exhaust memory "
        f"before any row could be processed. If the workbook is genuinely this "
        f"large, split it or export the sheet as .csv, which streams without a "
        f"limit."
    )


def _stream_xlsx_rows(workbook, row_iterator, headers: tuple[str, ...]):
    """Yield each remaining worksheet row as a header→text dict, then close the
    workbook (read-only workbooks hold an open zip handle until closed).

    All-empty rows are *buffered*, not skipped. "Trailing" is a property of
    position, so it cannot be decided by looking at a row's content: an
    all-empty row is padding only if nothing follows it. Skipping every blank
    row dropped an interior separator row that the CSV reader keeps, so the two
    readers disagreed on the row count for the same data — and output row N
    stopped corresponding to input row N, which silently shifts every district
    up by one for an operator pasting the pip_* columns back beside their sheet.
    Held blanks are released the moment a non-empty row arrives, and become
    ordinary rows the runner flags as errors; blanks still buffered at the end
    of the sheet are Excel's padding and are dropped.
    """
    # A blank row is fully described by its width, so the buffer holds widths
    # rather than the rows themselves: a sheet padded with a million blank rows
    # costs a list of a million integers, not a million lists of "".
    pending_blank_widths: list[int] = []
    try:
        for record in row_iterator:
            cells = _as_text_cells(record) or []
            if not any(cell for cell in cells):
                pending_blank_widths.append(len(cells))
                continue
            for width in pending_blank_widths:
                yield _row_from_cells(headers, [""] * width)
            pending_blank_widths.clear()
            yield _row_from_cells(headers, cells)
    finally:
        workbook.close()


def _read_google_sheet(
    url: str,
    *,
    sheet_gid: str | None,
    on_warning: Callable[[str], None] | None = None,
) -> Source:
    """Fetch one tab of a link-shared Google Sheet as CSV and parse it in memory.

    The whole body is read into memory (a Sheet is bounded by Google's own cell
    limit, and the export is a single non-resumable response) but is never
    written to disk — D15 allows no cache, no temp file.

    The failure this backend exists to explain: when the sheet is *not*
    link-shared, Google does not answer 403 with JSON. It answers 200 with an
    HTML sign-in page, which parses as perfectly valid one-column CSV full of
    markup. Left undetected that becomes thousands of nonsense rows, each one a
    geocoder call. We detect it and stop the run before a single address is sent.
    """
    doc_id, gid = parse_google_sheets_url(url, sheet_gid=sheet_gid)
    export_url = sheets_export_url(doc_id, gid)
    name = f"Google Sheet {doc_id} (gid {gid})"

    try:
        response = httpx.get(
            export_url, follow_redirects=True, timeout=SHEETS_TIMEOUT_SECONDS
        )
    except httpx.HTTPError as error:
        # Never let an httpx exception escape: the caller contract is BatchError.
        raise BatchError(
            f"could not reach Google Sheets for document {doc_id}: "
            f"{type(error).__name__}. Check the network connection and retry."
        ) from error

    if response.status_code >= 400:
        raise BatchError(
            f"Google Sheets returned HTTP {response.status_code} for document "
            f"{doc_id} (gid {gid}). This is a Google-side or URL problem, not a "
            f"sharing problem; verify the URL and retry."
        )

    body = response.text
    if _looks_like_sign_in_page(response, body):
        raise BatchError(
            f"Google Sheet {doc_id} is not link-shared: the export returned a "
            f"sign-in page instead of CSV. {SHARE_INSTRUCTIONS} "
            f"Then re-run. (Or download the sheet as .csv and pass the file.)"
        )

    reader = csv.reader(io.StringIO(body, newline=""))
    headers = _headers_from(next(reader, None), source_name=name)
    return Source(
        headers=headers,
        rows=(
            _row_from_cells(headers, record)
            for record in _records_warning_on_multiline_cells(
                reader, name, on_warning
            )
        ),
        name=name,
    )


def _looks_like_sign_in_page(response: httpx.Response, body: str) -> bool:
    """True when a Sheets export answered with Google's HTML login page.

    Three independent signals, because Google has changed which one it sends:
    an HTML content type, a body that opens as markup, or an accounts.google.com
    redirect landing in the body or the final URL.
    """
    content_type = response.headers.get("content-type", "").lower()
    if "text/html" in content_type:
        return True
    if body.lstrip()[:1] == "<":
        return True
    if "accounts.google.com" in str(response.url):
        return True
    return "accounts.google.com" in body[:4096]


# ---- shared helpers ----------------------------------------------------


def is_url(spec: str) -> bool:
    """True for an http(s) spec. Anything else is treated as a filesystem path
    (a Windows drive letter like ``C:\\data`` is not a scheme we accept).

    Public because `read_source`'s dispatch rule is also what a caller needs to
    know whether a source spec is a local file it can compare against an output
    path — the CLI's self-overwrite guard asks exactly that. One definition, so
    the guard can never disagree with the dispatch it is guarding.
    """
    return spec.strip().lower().startswith(("http://", "https://"))


def _first_query_value(query: str, key: str) -> str | None:
    """First value of `key` in a query- or fragment-encoded string, or None."""
    values = parse_qs(query).get(key)
    return values[0] if values else None


def _as_text_cells(record) -> list[str] | None:
    """Render a worksheet row's values as text. None (an empty cell) becomes "".

    A whole-number float is rendered without its ".0" tail: openpyxl reports the
    numeric cell 60602 as 60602.0, and a ZIP of "60602.0" geocodes to nothing.
    """
    if record is None:
        return None
    return [_cell_text(value) for value in record]


def _cell_text(value) -> str:
    """One cell value → text. Called from both `_as_text_cells` and
    `_row_from_cells`, so the whole-number-float rule lives here rather than at
    either caller: openpyxl reports the numeric cell 60602 as 60602.0, and a ZIP
    of "60602.0" geocodes to nothing."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _headers_from(cells: list[str] | None, *, source_name: str) -> tuple[str, ...]:
    """Validate and normalize the first row into the source's column names.

    Surrounding whitespace is trimmed — a header of "Address " is the operator's
    invisible typo, not a distinct column, and leaving it in makes
    `validate_column_mapping` reject a mapping that looks identical on screen.
    Nothing else about a header is changed.
    """
    if cells is None:
        raise BatchError(f"{source_name} is empty: no header row to read")

    headers = tuple(cell.strip() for cell in cells)
    # Trailing empty header cells are Excel's padding, not columns.
    while headers and headers[-1] == "":
        headers = headers[:-1]
    if not headers:
        raise BatchError(f"{source_name}: the header row is blank")

    blank_positions = [i + 1 for i, header in enumerate(headers) if header == ""]
    if blank_positions:
        raise BatchError(
            f"{source_name}: the header row has a blank column name at "
            f"position(s) {blank_positions}; every column needs a name"
        )

    duplicates = sorted({h for h in headers if headers.count(h) > 1})
    if duplicates:
        raise BatchError(
            f"{source_name}: duplicate column name(s) "
            f"{', '.join(repr(d) for d in duplicates)}; column names must be "
            f"unique so a mapping is unambiguous"
        )
    return headers


def _row_from_cells(headers: tuple[str, ...], cells) -> dict[str, str]:
    """Zip a record onto the headers. A short row pads with "", a long row drops
    its extra cells — neither may abort the run (plan D19)."""
    values = list(cells)
    if len(values) < len(headers):
        values.extend([""] * (len(headers) - len(values)))
    return {
        header: value if isinstance(value, str) else _cell_text(value)
        for header, value in zip(headers, values)
    }
